from openpyxl import load_workbook
# from random import *

wb = load_workbook("sample.xlsx")
ws = wb.active

# ws.append(["번호", "영어", "수학"])
# for i in range(1, 11): # 10개의 데이터 넣기
#     ws.append([1, randint(0,100), randint(0, 100)])

# ws.delete_cols(1,3) # 1번째 열로부터 3개 열 추가로 삭제

# find = input("v표를 누르세요 :")

# for x in range(1, ws.max_row + 1):
#     for y in range(1, ws.max_column + 1):
#         if (ws.cell(row=x, column=y).value == None): 
#             print(ws.cell(row=x, column=y).value, end=" ")
#         else:
            
        # print(ws.cell(row=x, column=y).value, end=" ") # end를 비게두면 1 2 3 4.. 로 나옴

# for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=242):
#     if row[3].value == None:
#         print(row[0].value, "호실 학생 검사요망")
#     # print()

# for x in range(1, 242):
#     for y in range(1, ws.max_column + 1):
#         # print(ws.cell(row=x, column=y).value, end =" ")
#         if ws.cell(row=x, column=4).value == None:
#             print(ws.cell(row=x, column=1).value, "호실 학생 검사 요망")
#         else:
#             print()
#     print()

# for x in range(1, ws.max_row + 1):
#     for y in range(1, ws.max_column + 1):
#         if ws.cell(row=x, column=2).value == 'v':
#             if ws.cell(row=x, column=3).value == None:
#                 if ws.cell(row=x, column=4).value == None:
#                     print(ws.cell(row=x, column=1).value, "호실 학생 검사 요망")
#     print()
