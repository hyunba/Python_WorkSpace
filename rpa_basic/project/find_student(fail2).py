from openpyxl import load_workbook
wb = load_workbook("20210923.xlsx")
ws = wb.active

# for row in ws.iter_rows(min_row=2, max_row=4, min_col=4, max_col=6):
#     print(row[4].value)


# for row in ws.iter_rows(min_row=1, max_row=3, min_col=3, max_col=8):
#     if row[10].value == None:
#         print(row[2].value, "호실 학생 검사요망")

# rows = ws.rows

# for row in rows:
#     for cell in row:
# 	        print(cell.value)

# for x in range(1, 1216):
#     for y in range(1, ws.max_column + 1):
#         print(ws.cell(row=x, column=y).value, end =" ")
#         if ws.cell(row=x, column=13).value == None:
#             print(ws.cell(row=x, column=6).value, "호실 학생 검사 요망")
#     print()

# for row in ws.iter_rows(min_row=1, max_row=13, min_col=1, max_col=1216):
#     if ws.cell(row).value == None:
#         print(row[6].value, "호실 학생 검사요망")

# wb.save("20210923_save_point.xlsx")
