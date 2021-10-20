from openpyxl import load_workbook

wb = load_workbook("sample_20211019.xlsx")
ws = wb.active


for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        if ws.cell(row=x, column=2).value == 'v':
            if ws.cell(row=x, column=3).value == None:
                if ws.cell(row=x, column=4).value == None:
                    print(ws.cell(row=x, column=1).value, "")
    print()
