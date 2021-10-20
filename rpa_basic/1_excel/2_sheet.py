from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성
ws.title="MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "ffbbff" # RGB 형태로 값을 넣어주면 탭 색상 변경

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 sheet 생성

new_ws = wb["NewSheet"] # Dick 형태로 Sheet에 접근 가능

print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws) # NewSheet의 A1에 있는 "Test"를 복사해서 Copied Sheet를 생성
target.title = "Copied Sheet"

wb.save("sample.xlsx")