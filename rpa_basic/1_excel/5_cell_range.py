from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# # 1줄씩 데이터 넣기
# ws.append(["번호", "영어", "수학"])
# for i in range(1, 11): # 10개의 데이터 넣기
#     ws.append([1, randint(0,100), randint(0, 100)])

# col_B = ws["B"] # 영어 column만 가져오기
# # print(col_B)
# # for cell in col_B:
# #     print(cell.value)

# col_range = ws["B:C"] # B~C까지의 column 값을 가져오기
# # for cols in col_range:
# #     for cell in cols:

# row_title = ws[1] # 1번째 row 만 가지고 오기
# # for cell in row_title:
# #     print(cell.value)

# # row_range = ws[2:6] # 1번째를 제외 2번째 줄에서 6번째 줄까지 가지고 오기
# # for rows in row_range:
# #     for cell in rows:
# #         print(cell.value, end= " ")
# #     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end= " ") # 셀의 좌표정보를 가져온다
#         xy = coordinate_from_string(cell.coordinate)
#         print(xy, end=" ")
#     print()

# 전체 rows
print(tuple(ws.rows))

# 전체 columns
print(tuple(ws.columns))

wb.save("sample.xlsx")