from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 셀에 1 이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 '값'을 출력

# row = 1, 2, 3, ...
# colume = A(1), B(2), C(3), ...
print(ws.cell(row=1, column=1).value) # ws["A1"].value

c = ws.cell(row=1, column=3, value=10) # ws["C1"].value = 10
print(c.value)

from random import *

# 반복문을 사용하기
index = 0
for x in range(1, 11): # 10개의 row
    for y in range(1, 11): # 10개의 column
        ws.cell(row=x, column=y, value=index) # index를 넣어서 0부터 계속 증가
        index += 1

wb.save("sample.xlsx")
